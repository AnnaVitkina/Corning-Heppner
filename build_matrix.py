"""Convert processed tariff tabs into wide matrix-form shipment and cost layout."""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

_colab_code = Path("/content/Corning-Heppner")
if _colab_code.exists() and str(_colab_code) not in sys.path:
    sys.path.insert(0, str(_colab_code))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from paths import get_paths

EXPECTED_ORIGIN_COUNTRY = "FR"
EXPECTED_ORIGIN_POSTAL_CODE = "44160"
EXPECTED_CLIENT_NAME = "CORNING POUYET"
EXPECTED_ACCOUNT = "767534"
COST_NAME = "Transport cost"
CURRENCY = "EUR"
SERVICE = "Standard"
FLAT_RULE = "Flat rule"

SHIPMENT_HEADERS = [
    "Lane #",
    "Origin Country",
    "Origin Postal Code",
    "Destination Country",
    "Destination Postal Code",
    "Service",
]

DEPT_MARKERS = ("dpt", "dept", "dépt")
LIBELLE_MARKERS = ("libell", "libelle")
ORIGIN_LOCATION_PATTERN = re.compile(
    r"(?P<country>[A-Z]{2})\s*-\s*(?P<postal>\d{4,5})",
    re.IGNORECASE,
)
DESTINATION_COUNTRY_PATTERN = re.compile(r"Tarif\s+([A-Z]{2})\b", re.IGNORECASE)
WEIGHT_NUMBER_PATTERN = re.compile(r"(\d+(?:\.\d+)?)")
INVALID_ROW_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
HEADER_FILL = PatternFill(start_color="FFB4C6E7", end_color="FFB4C6E7", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FF000000")
THIN_BORDER = Border(
    left=Side(style="thin", color="FFBFBFBF"),
    right=Side(style="thin", color="FFBFBFBF"),
    top=Side(style="thin", color="FFBFBFBF"),
    bottom=Side(style="thin", color="FFBFBFBF"),
)

HEADER_ROW_COUNT = 5
DATA_START_ROW = HEADER_ROW_COUNT + 1


@dataclass
class LaneRow:
    origin_country: str
    origin_postal_code: str
    destination_country: str
    destination_postal_code: str
    costs: dict[str, object]


@dataclass
class MatrixSheet:
    lanes: list[LaneRow]
    bracket_headers: list[str]
    rate_by: str
    cost_title: str
    origin_valid: bool


def list_processed_files() -> list[Path]:
    processing_dir = get_paths().processing_dir
    return sorted(
        path
        for path in processing_dir.glob("*_processed.xlsx")
        if not path.name.startswith("~$")
    )


def prompt_file_selection(files: list[Path]) -> list[Path]:
    print("\nProcessed files:")
    for index, file_path in enumerate(files, start=1):
        print(f"  {index}. {file_path.name}")

    print("\nEnter file numbers to convert (comma-separated), e.g. 1,2")
    print("Or enter 'all' to convert every processed file.")
    choice = input("Your choice: ").strip().lower()

    if choice == "all":
        return files

    selected_indexes: set[int] = set()
    for part in choice.split(","):
        part = part.strip()
        if not part:
            continue
        if not part.isdigit():
            raise ValueError(f"Invalid selection: {part!r}")
        selected_indexes.add(int(part))

    selected_files = []
    for index in sorted(selected_indexes):
        if index < 1 or index > len(files):
            raise ValueError(f"Selection out of range: {index}")
        selected_files.append(files[index - 1])

    if not selected_files:
        raise ValueError("No files selected.")

    return selected_files


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def preserve_header_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    if "\n" in text:
        return "\n".join(part.strip() for part in text.splitlines() if part.strip())
    return text.strip()


def is_dept_header(value: object) -> bool:
    text = normalize_text(value).lower()
    return any(marker in text for marker in DEPT_MARKERS)


def is_libelle_header(value: object) -> bool:
    text = normalize_text(value).lower()
    return any(marker in text for marker in LIBELLE_MARKERS)


def find_table_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if is_libelle_header(ws.cell(row, col).value):
                return row
    return None


def find_column_by_header(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    matcher,
) -> int | None:
    for col in range(1, ws.max_column + 1):
        if matcher(ws.cell(header_row, col).value):
            return col
    return None


def read_preamble_value(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int = 17) -> str:
    return normalize_text(ws.cell(row, col).value)


def parse_origin_location(location_text: str) -> tuple[str, str]:
    match = ORIGIN_LOCATION_PATTERN.search(location_text)
    if not match:
        return "", ""
    return match.group("country").upper(), match.group("postal")


def validate_origin(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[bool, str, str]:
    client_name = read_preamble_value(ws, 1)
    location_text = read_preamble_value(ws, 2)
    account_text = read_preamble_value(ws, 3)

    origin_country, origin_postal_code = parse_origin_location(location_text)

    is_valid = (
        client_name.upper() == EXPECTED_CLIENT_NAME
        and origin_country == EXPECTED_ORIGIN_COUNTRY
        and origin_postal_code == EXPECTED_ORIGIN_POSTAL_CODE
        and EXPECTED_ACCOUNT in account_text
    )

    return is_valid, origin_country, origin_postal_code


def parse_destination_country(sheet_name: str) -> str:
    match = DESTINATION_COUNTRY_PATTERN.search(sheet_name)
    if not match:
        raise ValueError(f"Could not parse destination country from sheet name: {sheet_name}")
    return match.group(1).upper()


def parse_rate_by(ws: openpyxl.worksheet.worksheet.Worksheet) -> str:
    for row in range(1, 12):
        value = ws.cell(row, 2).value
        if value is None:
            continue
        text = normalize_text(value).lower()
        if "plancher" in text:
            return "Area/ldm"
        if "poids" in text:
            return "Weight"
    return "Weight"


def parse_weight_bracket(header: object) -> tuple[str, str]:
    raw_label = preserve_header_text(header)
    collapsed = normalize_text(header)

    numbers = WEIGHT_NUMBER_PATTERN.findall(collapsed)
    if not numbers:
        return "", raw_label or collapsed

    if len(numbers) >= 2:
        lower, upper = numbers[0], numbers[-1]
        label = f"{lower} - {upper}"
        return f"<= {upper}", label

    upper = numbers[0]
    return f"<= {upper}", upper


def bracket_sort_key(apply_if_header: str) -> float:
    match = re.search(r"<=\s*(\d+(?:\.\d+)?)", apply_if_header)
    if not match:
        return 0.0
    return float(match.group(1))


def bracket_bounds(label: str) -> tuple[str, str]:
    numbers = WEIGHT_NUMBER_PATTERN.findall(label)
    if not numbers:
        return label, label
    if len(numbers) == 1:
        return numbers[0], numbers[0]
    return numbers[0], numbers[-1]


def build_cost_title(bracket_labels: list[str], rate_by: str) -> str:
    if not bracket_labels:
        return COST_NAME

    low, _ = bracket_bounds(bracket_labels[0])
    _, high = bracket_bounds(bracket_labels[-1])
    unit = "pallet plancher" if rate_by == "Area/ldm" else "weight"
    return f"{COST_NAME} ({low} to {high} {unit})"


def get_cost_columns(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    libelle_col: int,
) -> list[tuple[int, str, str]]:
    cost_columns: list[tuple[int, str, str]] = []

    for col in range(libelle_col + 1, ws.max_column + 1):
        header_value = ws.cell(header_row, col).value
        if header_value in (None, ""):
            continue

        apply_if, weight_bracket = parse_weight_bracket(header_value)
        if not apply_if:
            continue

        cost_columns.append((col, apply_if, weight_bracket))

    if not cost_columns:
        raise ValueError("No weight bracket cost columns found.")

    cost_columns.sort(key=lambda item: bracket_sort_key(item[1]))
    return cost_columns


def build_matrix(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    sheet_name: str,
) -> MatrixSheet:
    header_row = find_table_header_row(ws)
    if header_row is None:
        raise ValueError(f"Could not find data table in sheet: {sheet_name}")

    dept_col = find_column_by_header(ws, header_row, is_dept_header)
    libelle_col = find_column_by_header(ws, header_row, is_libelle_header)
    if dept_col is None:
        raise ValueError(f"Could not find Dépt column in sheet: {sheet_name}")
    if libelle_col is None:
        raise ValueError(f"Could not find Libellé column in sheet: {sheet_name}")

    origin_valid, parsed_country, parsed_postal = validate_origin(ws)
    destination_country = parse_destination_country(sheet_name)
    rate_by = parse_rate_by(ws)
    cost_columns = get_cost_columns(ws, header_row, libelle_col)
    bracket_headers = [apply_if for _, apply_if, _ in cost_columns]
    bracket_labels = [label for _, _, label in cost_columns]
    cost_title = build_cost_title(bracket_labels, rate_by)

    lanes: list[LaneRow] = []
    for row in range(header_row + 1, ws.max_row + 1):
        destination_postal_code = normalize_text(ws.cell(row, dept_col).value)
        if not destination_postal_code:
            if lanes:
                break
            continue

        costs: dict[str, object] = {}
        for cost_col, apply_if, _ in cost_columns:
            transport_cost = ws.cell(row, cost_col).value
            if transport_cost not in (None, ""):
                costs[apply_if] = transport_cost

        lanes.append(
            LaneRow(
                origin_country=EXPECTED_ORIGIN_COUNTRY,
                origin_postal_code=EXPECTED_ORIGIN_POSTAL_CODE,
                destination_country=destination_country,
                destination_postal_code=destination_postal_code,
                costs=costs,
            )
        )

    if not lanes:
        raise ValueError(f"No matrix rows found in sheet: {sheet_name}")

    if not origin_valid:
        print(
            f"  ! Origin validation failed for {sheet_name}: "
            f"expected {EXPECTED_ORIGIN_COUNTRY}-{EXPECTED_ORIGIN_POSTAL_CODE}, "
            f"found {parsed_country}-{parsed_postal or 'N/A'}"
        )

    return MatrixSheet(
        lanes=lanes,
        bracket_headers=bracket_headers,
        rate_by=rate_by,
        cost_title=cost_title,
        origin_valid=origin_valid,
    )


def style_cell(
    cell: openpyxl.cell.cell.Cell,
    *,
    header: bool = False,
    horizontal: str = "center",
) -> None:
    if header:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def merge_and_set(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
    value: str,
) -> None:
    if min_row != max_row or min_col != max_col:
        worksheet.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )
    cell = worksheet.cell(row=min_row, column=min_col, value=value)
    style_cell(cell, header=True)


def write_matrix_sheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    matrix: MatrixSheet,
) -> None:
    shipment_col_count = len(SHIPMENT_HEADERS)
    currency_col = shipment_col_count + 1
    first_bracket_col = currency_col + 1
    last_col = first_bracket_col + len(matrix.bracket_headers) - 1

    # Shipment headers merged vertically across all header rows.
    for col_idx, header in enumerate(SHIPMENT_HEADERS, start=1):
        merge_and_set(
            worksheet,
            min_row=1,
            min_col=col_idx,
            max_row=HEADER_ROW_COUNT,
            max_col=col_idx,
            value=header,
        )

    # Transport cost block: merged title / rate by / flat rule across Currency + brackets.
    merge_and_set(
        worksheet,
        min_row=1,
        min_col=currency_col,
        max_row=1,
        max_col=last_col,
        value=matrix.cost_title,
    )
    merge_and_set(
        worksheet,
        min_row=2,
        min_col=currency_col,
        max_row=2,
        max_col=last_col,
        value=f"Rate by: {matrix.rate_by}",
    )
    merge_and_set(
        worksheet,
        min_row=3,
        min_col=currency_col,
        max_row=3,
        max_col=last_col,
        value=FLAT_RULE,
    )

    currency_cell = worksheet.cell(row=4, column=currency_col, value="Currency")
    style_cell(currency_cell, header=True)

    for offset, bracket_header in enumerate(matrix.bracket_headers):
        bracket_col = first_bracket_col + offset
        bracket_cell = worksheet.cell(row=4, column=bracket_col, value=bracket_header)
        style_cell(bracket_cell, header=True)

        flat_cell = worksheet.cell(row=5, column=bracket_col, value="Flat")
        style_cell(flat_cell, header=True)

    # Row 5 under Currency stays blank (only bracket columns show "Flat").
    style_cell(worksheet.cell(row=5, column=currency_col), header=True)

    for lane_index, lane in enumerate(matrix.lanes, start=1):
        row_idx = DATA_START_ROW + lane_index - 1
        row_values = [
            lane_index,
            lane.origin_country,
            lane.origin_postal_code,
            lane.destination_country,
            lane.destination_postal_code,
            SERVICE,
            CURRENCY,
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            style_cell(cell, horizontal="left" if col_idx in {2, 3, 4, 5, 6, 7} else "center")

        for offset, bracket_header in enumerate(matrix.bracket_headers):
            value = lane.costs.get(bracket_header)
            cell = worksheet.cell(row=row_idx, column=first_bracket_col + offset, value=value)
            style_cell(cell, horizontal="right")

        if not matrix.origin_valid:
            for col_idx in range(1, last_col + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = INVALID_ROW_FILL

    worksheet.freeze_panes = worksheet.cell(row=DATA_START_ROW, column=1)
    worksheet.auto_filter.ref = (
        f"A{HEADER_ROW_COUNT}:{get_column_letter(last_col)}{HEADER_ROW_COUNT}"
    )

    for col_idx in range(1, shipment_col_count + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 18
    worksheet.column_dimensions[get_column_letter(currency_col)].width = 12
    for col_idx in range(first_bracket_col, last_col + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 11


def convert_file(file_path: Path) -> Path:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    matrix_tables: dict[str, MatrixSheet] = {}

    for sheet_name in workbook.sheetnames:
        matrix = build_matrix(workbook[sheet_name], sheet_name)
        matrix_tables[sheet_name] = matrix
        status = "valid origin" if matrix.origin_valid else "INVALID origin (rows highlighted)"
        print(
            f"  - {sheet_name}: {len(matrix.lanes)} lanes x "
            f"{len(matrix.bracket_headers)} brackets, {status}"
        )

    workbook.close()

    output_path = get_paths().output_dir / f"{file_path.stem.replace('_processed', '')}_matrix.xlsx"
    output_workbook = openpyxl.Workbook()
    output_workbook.remove(output_workbook.active)

    for sheet_name, matrix in matrix_tables.items():
        worksheet = output_workbook.create_sheet(title=sheet_name[:31])
        write_matrix_sheet(worksheet, matrix)

    output_workbook.save(output_path)
    output_workbook.close()

    return output_path


def main() -> int:
    paths = get_paths()
    processing_dir = paths.processing_dir
    output_dir = paths.output_dir

    if not processing_dir.exists():
        print(f"Processing folder not found: {processing_dir}")
        return 1

    output_dir.mkdir(parents=True, exist_ok=True)

    files = list_processed_files()
    if not files:
        print(f"No processed files found in {processing_dir}")
        print("Run process_input.py first.")
        return 1

    print(f"Environment: {paths.environment}")
    print(f"Data root:     {paths.data_root}")

    try:
        selected_files = prompt_file_selection(files)
    except ValueError as exc:
        print(f"Error: {exc}")
        return 1

    print(f"\nBuilding matrix for {len(selected_files)} file(s)...")
    for file_path in selected_files:
        print(f"\n{file_path.name}")
        output_path = convert_file(file_path)
        print(f"Saved: {output_path}")

    print("\nDone.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
