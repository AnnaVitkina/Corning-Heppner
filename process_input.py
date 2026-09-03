"""Read Excel files from input/, prompt user in terminal, save cleaned tabs to processing/."""

from __future__ import annotations

import re
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

_colab_code = Path("/content/Corning-Heppner")
if _colab_code.exists() and str(_colab_code) not in sys.path:
    sys.path.insert(0, str(_colab_code))

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment

from paths import exit_with_code, get_paths

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
EXCLUDED_NAME_ENDINGS = (
    "_processed.xlsx",
    "_processed.xlsm",
    "_processed.xls",
    "_matrix.xlsx",
    "_matrix.xlsm",
    "_shipment_matrix.xlsx",
    "_shipment_matrix.xlsm",
)

LIBELLE_MARKERS = ("libell", "libelle")
FOOTNOTE_MARKERS = ("poids arrondi", "tarif ht en euro")


def is_source_input_file(path: Path) -> bool:
    if not path.is_file() or path.name.startswith("~$"):
        return False
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        return False
    name_lower = path.name.lower()
    return not any(name_lower.endswith(ending) for ending in EXCLUDED_NAME_ENDINGS)


def list_input_files() -> list[Path]:
    input_dir = get_paths().input_dir
    files: list[Path] = []
    skipped: list[str] = []

    for path in sorted(input_dir.iterdir()):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        if is_source_input_file(path):
            files.append(path)
        else:
            skipped.append(path.name)

    if skipped:
        print("\nSkipped files in input/ (not original supplier tariffs):")
        for name in skipped:
            print(f"  - {name}")
        print("Only place the original Heppner .xlsx files in input/.")

    return files


def validate_source_workbook(file_path: Path) -> None:
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            if find_header_row(workbook[sheet_name]) is not None:
                return
    finally:
        workbook.close()

    raise ValueError(
        f"{file_path.name} does not look like a source Heppner tariff file.\n"
        "Expected tabs with a 'Libellé' column.\n"
        "Put the original supplier workbook in input/, not *_processed.xlsx or *_matrix.xlsx."
    )


def prompt_file_selection(files: list[Path]) -> list[Path]:
    print("\nFiles in input folder:")
    for index, file_path in enumerate(files, start=1):
        print(f"  {index}. {file_path.name}")

    print("\nEnter file numbers to process (comma-separated), e.g. 1,2")
    print("Or enter 'all' to process every file.")
    choice = input("Your choice: ").strip().lower()

    if choice == "all":
        return files

    selected_indexes: set[int] = set()
    for part in choice.split(","):
        part = part.strip()
        if not part:
            continue
        if not part.isdigit():
            raise ValueError(f"Invalid selection: {part!r}")
        selected_indexes.add(int(part))

    selected_files: list[Path] = []
    for index in sorted(selected_indexes):
        if index < 1 or index > len(files):
            raise ValueError(f"Selection out of range: {index}")
        selected_files.append(files[index - 1])

    if not selected_files:
        raise ValueError("No files selected.")

    return selected_files


def sanitize_sheet_name(name: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name).strip() or "Sheet"
    cleaned = cleaned[:31]

    candidate = cleaned
    suffix = 1
    while candidate in used_names:
        tail = f"_{suffix}"
        candidate = f"{cleaned[: 31 - len(tail)]}{tail}"
        suffix += 1

    used_names.add(candidate)
    return candidate


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def preserve_header_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    if "\n" in text:
        return "\n".join(part.strip() for part in text.splitlines() if part.strip())
    return text.strip()


def is_libelle_header(value: object) -> bool:
    text = normalize_text(value).lower()
    return any(marker in text for marker in LIBELLE_MARKERS)


def is_footnote_row(libelle_value: object, dept_value: object) -> bool:
    libelle_text = normalize_text(libelle_value).lower()
    dept_text = normalize_text(dept_value).lower()

    if not libelle_text:
        return bool(dept_text and len(dept_text) > 20)

    return any(marker in libelle_text for marker in FOOTNOTE_MARKERS)


def format_display_value(value: object, number_format: str | None) -> object:
    if value is None:
        return None

    if not isinstance(value, (int, float)):
        return value

    fmt = (number_format or "General").split(";")[0].strip()

    if fmt in {"General", "@", ""}:
        return value

    if "0.00" in fmt or "#,##0.00" in fmt:
        decimals = len(fmt.split(".")[-1]) if "." in fmt else 2
        quant = Decimal("1").scaleb(-decimals)
        rounded = Decimal(str(value)).quantize(quant, rounding=ROUND_HALF_UP)
        formatted = f"{rounded:.{decimals}f}"
        if "#,##" in fmt:
            whole, fraction = formatted.split(".")
            sign = "-" if whole.startswith("-") else ""
            whole = whole.lstrip("-")
            whole = f"{int(whole):,}"
            formatted = f"{sign}{whole}.{fraction}"
        return formatted

    if fmt in {"0", "#,##0"} or fmt.startswith("#,##0"):
        rounded = int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        if "#,##" in fmt:
            return f"{rounded:,}"
        return str(rounded)

    if "0.0" in fmt:
        decimals = len(fmt.split(".")[-1])
        quant = Decimal("1").scaleb(-decimals)
        rounded = Decimal(str(value)).quantize(quant, rounding=ROUND_HALF_UP)
        return f"{rounded:.{decimals}f}"

    return value


def find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if is_libelle_header(ws.cell(row, col).value):
                return row
    return None


def is_zero_placeholder_row(ws: openpyxl.worksheet.worksheet.Worksheet, row: int) -> bool:
    values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
    non_empty = [value for value in values if value not in (None, "")]
    return bool(non_empty) and all(value == 0 for value in non_empty)


def extract_preamble(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    max_col: int = 17,
) -> list[list[object | None]]:
    preamble: list[list[object | None]] = []

    for row in range(1, header_row):
        if is_zero_placeholder_row(ws, row):
            continue

        row_values: list[object | None] = []
        has_content = False
        for col in range(1, max_col + 1):
            value = ws.cell(row, col).value
            if value not in (None, ""):
                has_content = True
            row_values.append(value)

        if has_content:
            preamble.append(row_values)

    return preamble


def write_preamble(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    preamble: list[list[object | None]],
) -> int:
    for row_idx, row_values in enumerate(preamble, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            if value not in (None, ""):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

    return len(preamble)


def find_last_data_column(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int) -> int:
    last_col = 1
    for col in range(1, ws.max_column + 1):
        if ws.cell(header_row, col).value is not None:
            last_col = col
    return last_col


def clean_header_label(value: object) -> str:
    text = preserve_header_text(value)
    return text if text else "Column"


def extract_clean_sheet(
    ws_values: openpyxl.worksheet.worksheet.Worksheet,
    ws_formats: openpyxl.worksheet.worksheet.Worksheet,
) -> tuple[list[list[object | None]], pd.DataFrame]:
    header_row = find_header_row(ws_values)
    if header_row is None:
        raise ValueError("Could not find table header row with 'Libellé'.")

    preamble = extract_preamble(ws_values, header_row)
    last_col = find_last_data_column(ws_values, header_row)
    libelle_col = next(
        (
            col
            for col in range(1, last_col + 1)
            if is_libelle_header(ws_values.cell(header_row, col).value)
        ),
        3,
    )
    dept_col = libelle_col - 1 if libelle_col > 1 else 2

    first_cost_col = libelle_col + 1

    headers: list[str] = []
    used_headers: set[str] = set()
    for col in range(dept_col, last_col + 1):
        label = clean_header_label(ws_values.cell(header_row, col).value)
        candidate = label
        suffix = 2
        while candidate in used_headers:
            candidate = f"{label}_{suffix}"
            suffix += 1
        used_headers.add(candidate)
        headers.append(candidate)

    rows: list[list[object]] = []
    for row in range(header_row + 1, ws_values.max_row + 1):
        libelle_value = ws_values.cell(row, libelle_col).value
        dept_value = ws_values.cell(row, dept_col).value

        if is_footnote_row(libelle_value, dept_value):
            break

        if libelle_value is None and dept_value is None:
            if rows:
                break
            continue

        if libelle_value is None:
            continue

        row_values: list[object] = []
        for offset, col in enumerate(range(dept_col, last_col + 1)):
            value = ws_values.cell(row, col).value
            if col >= first_cost_col:
                number_format = ws_formats.cell(row, col).number_format
                value = format_display_value(value, number_format)
            elif col == libelle_col:
                value = normalize_text(value)
            else:
                value = normalize_text(value) if value is not None else value
            row_values.append(value)

        rows.append(row_values)

    return preamble, pd.DataFrame(rows, columns=headers)


def process_file(file_path: Path) -> Path:
    validate_source_workbook(file_path)
    wb_values = openpyxl.load_workbook(file_path, data_only=True)
    wb_formats = openpyxl.load_workbook(file_path, data_only=False)

    sheet_outputs: dict[str, tuple[list[list[object | None]], pd.DataFrame]] = {}
    used_sheet_names: set[str] = set()

    for sheet_name in wb_values.sheetnames:
        preamble, df = extract_clean_sheet(wb_values[sheet_name], wb_formats[sheet_name])
        output_sheet_name = sanitize_sheet_name(sheet_name, used_sheet_names)
        sheet_outputs[output_sheet_name] = (preamble, df)
        print(
            f"  - {sheet_name}: {len(preamble)} header rows, "
            f"{df.shape[0]} data rows x {df.shape[1]} columns"
        )

    wb_values.close()
    wb_formats.close()

    output_path = get_paths().processing_dir / f"{file_path.stem}_processed.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, (preamble, df) in sheet_outputs.items():
            table_start_row = len(preamble) + 1
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=table_start_row,
            )
            worksheet = writer.sheets[sheet_name]
            write_preamble(worksheet, preamble)
            header_row_idx = table_start_row + 1
            for col_idx, column_name in enumerate(df.columns, start=1):
                header_cell = worksheet.cell(row=header_row_idx, column=col_idx)
                if isinstance(column_name, str) and "\n" in column_name:
                    header_cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="center",
                        horizontal="center",
                    )

    return output_path


def main() -> int:
    paths = get_paths()
    input_dir = paths.input_dir
    processing_dir = paths.processing_dir

    if not input_dir.exists():
        print(f"Input folder not found: {input_dir}")
        return 1

    processing_dir.mkdir(parents=True, exist_ok=True)

    files = list_input_files()
    if not files:
        print(f"No source tariff files found in {input_dir}")
        print("Add the original Heppner supplier .xlsx to input/.")
        print("Do not use *_processed.xlsx or *_matrix.xlsx files as input.")
        return 1

    print(f"Environment: {paths.environment}")
    print(f"Data root:     {paths.data_root}")

    try:
        selected_files = prompt_file_selection(files)
    except ValueError as exc:
        print(f"Error: {exc}")
        return 1

    print(f"\nProcessing {len(selected_files)} file(s)...")
    for file_path in selected_files:
        print(f"\n{file_path.name}")
        output_path = process_file(file_path)
        print(f"Saved: {output_path}")

    print("\nDone.")
    return 0


if __name__ == "__main__":
    exit_with_code(main())
